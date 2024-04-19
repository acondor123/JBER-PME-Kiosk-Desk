from openpyxl import Workbook
from openpyxl import load_workbook
from pyzbar.pyzbar import decode
from PIL import Image


def write_to_excel(decodeQR):
    try:
        try:
            workbook = load_workbook('entries.xlsx')
        except FileNotFoundError:
            workbook = Workbook()
            workbook.save('entries.xlsx')

        worksheet = workbook.active

        # Find the next available row
        next_row = worksheet.max_row + 1

        # Split the string and write it to the columns
        qr_data = decodeQR.split(',')
        for i, data in enumerate(qr_data):
            worksheet.cell(row=next_row, column=i + 1).value = data
        next_row += 1

        # Save workbook
        workbook.save('entries.xlsx')

        return True
    except Exception as error:
        print(f"Error writing QR code data to spreadsheet: {error}")
        return False

# Testing with existing QR code, still need to take in active time
decodeQR = decode(Image.open('qr_code.png'))
decodeQR2 = decode(Image.open('qr_code.png'))
entry = decodeQR[0].data.decode('utf-8')
entry2 = decodeQR2[0].data.decode('utf-8')

write_to_excel(entry)
write_to_excel(entry2)

