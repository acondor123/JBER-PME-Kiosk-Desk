import sys
import time
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QLabel, QDesktopWidget, QMessageBox
from PyQt5.QtGui import QColor, QPixmap, QMovie
from PyQt5.QtCore import Qt, QTimer
from Resources.validate import *
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

'''
    This file contains the source code for the Kiosk app.
    All required installations are shown in requirements.txt.
    The validate.py import is located in the resources folder.

    For kiosk.py
        Connor wrote    70%
        Mika wrote      30%

    For validate.py
        Connor wrote    100%
        Mika wrote      0%
'''

class QRCodeScanner(QWidget):
    def __init__(self):
        super().__init__()

        self.scanned_code = ""
        self.currently_scanning = False
        self.data_fields = {
            "first_name": False,
            "last_name": False,
            "rank": False,
            "unit": False,
            "phone_number": False,
            "fitness": False,
            "profile": False
        }

        self.timer = QTimer()
        self.timer.timeout.connect(self.timer_expired)

        self.setWindowTitle("QR Code Scanner")

        screen_geometry = QDesktopWidget().screenGeometry()

        width = int(screen_geometry.width() * 0.8)
        height = int(screen_geometry.height() * 0.6)

        layout = QVBoxLayout(self)

        outer_container = QWidget()
        outer_container.setAutoFillBackground(True)
        outer_palette = outer_container.palette()
        outer_palette.setColor(outer_container.backgroundRole(), QColor("#D3D3D3"))
        outer_container.setPalette(outer_palette)

        outer_container_layout = QVBoxLayout(outer_container)

        inner_container = QWidget()
        inner_container.setAutoFillBackground(True)
        inner_palette = inner_container.palette()
        inner_palette.setColor(inner_container.backgroundRole(), QColor("#ffffff"))
        inner_container.setPalette(inner_palette)

        inner_container.setStyleSheet("background-color: #ffffff; border-radius: 10px;")

        inner_container_layout = QVBoxLayout(inner_container)

        inner_container.setFixedWidth(int(width / 1.1))
        inner_container.setFixedHeight(int(height * 1.2))

        message_label = QLabel(inner_container)
        message_label.setText("Please scan your QR code")
        message_label.setStyleSheet("font-size: 48px; color: #333; font-family: Arial, sans-serif;")
        message_label.setAlignment(Qt.AlignCenter)

        image_label = QLabel(inner_container)
        pixmap = QPixmap("images/pmelogo.png")
        image_label.setPixmap(pixmap)
        image_label.setAlignment(Qt.AlignCenter)

        inner_container_layout.addWidget(message_label, alignment=Qt.AlignCenter)
        inner_container_layout.addWidget(image_label, alignment=Qt.AlignCenter)

        outer_container_layout.addWidget(inner_container, alignment=Qt.AlignCenter)

        layout.addWidget(outer_container)

        self.loading_container = QWidget(self)
        self.loading_container.setGeometry(0, 0, screen_geometry.width(), screen_geometry.height())
        self.loading_container.setStyleSheet("background-color: rgba(255, 255, 255, 150);")
        self.loading_layout = QVBoxLayout(self.loading_container)
        self.loading_label = QLabel()
        self.loading_layout.addWidget(self.loading_label, alignment=Qt.AlignCenter)
        self.loading_container.hide()

    '''
        Below function sets GUI to fullscreen when double tapped/clicked.
    '''

    def mouseDoubleClickEvent(self, event):
        if event.button() == Qt.LeftButton:
            if self.isFullScreen():
                self.showNormal()
            else:
                self.showFullScreen()

    '''
        Below funciton deals with the logic for when a qrcode is scanned.
        Since barcode scanners are treated as keyboard input, we have
        decided to use keypressevents to determine what has been entered
    '''
    def keyPressEvent(self, event):
        #for each button press, add each letter to the scanned_code variable
        if event.text() != '$' and event.text() != "":
            self.scanned_code += event.text()
            if not self.currently_scanning:
                self.load_screen()
                self.currently_scanning = True
                #set timer to make sure the machine resets if an invalid qrcode is scanned that doesn't
                #have the terminal signal in it; if timer expires, calls reset_scanned_code()
                self.timer.start(5000)
        
        #now that the terminal symbol is reached, we need to validate the input and reset the machine's state for the next scan
        elif self.scanned_code != "" and event.text() == "$":
            print(self.scanned_code)
            is_valid = self.validate_input(self.scanned_code)
            if(is_valid):
                self.update_spreadsheet()
                self.reset_scanned_code()
                self.show_checkmark_overlay()            
            else:
                self.reset_scanned_code()
                self.show_invalid_qr_message()
            time.sleep(0.5)

    '''
        Below function sets the machine's state to its initial state to ensure no bugs are introduced.
    '''
    def reset_scanned_code(self):
        self.scanned_code = ""
        self.load_screen()
        self.currently_scanning = False
        self.timer.stop()
        self.data_fields = {
            "first_name": False,
            "last_name": False,
            "rank": False,
            "unit": False,
            "phone_number": False,
            "fitness": False,
            "profile": False
        }

    '''
        Below function specifies what happens when a scan of a QR code takes too long.
        This indicates that an invalid QR code is scanned.
    '''

    def timer_expired(self):
        self.reset_scanned_code()
        self.show_invalid_qr_message()

    '''
        Below funciton creates a new spreadsheet with the information scanned.  If the
        workbook already exists, it simply adds on to it.
    '''

    def update_spreadsheet(self):
        current_date = datetime.date.today()
        formatted_date = current_date.strftime("%d-%b-%Y")
        try:
            try:
                workbook = load_workbook(f'spreadsheets/{formatted_date}.xlsx')
            except FileNotFoundError:
                workbook = Workbook()
                workbook.save(f'spreadsheets/{formatted_date}.xlsx')

            worksheet = workbook.active

            next_row = worksheet.max_row + 1

            qr_data = self.scanned_code.split(',')
            for i, data in enumerate(qr_data):
                worksheet.cell(row=next_row, column=i + 1).value = data
            next_row += 1

            workbook.save(f'spreadsheets/{formatted_date}.xlsx')

            print("QR code data written to spreadsheet successfully")
        except Exception as error:
            print(f"Error writing QR code data to spreadsheet: {error}")

    '''
        Below function validates the user's input to ensure it is in the same format as given to them in the website.
        This ensures nothing malicious is scanned, as it will reject anything that isn't expected.
    '''
    def validate_input(self, qr_data):
        valid_length = 7
        qr_data = qr_data.split(",")

        if(len(qr_data) > valid_length):
            print("Too much data")
            return False
        elif(len(qr_data) < valid_length):
            print("Not enough data")
            return False
        #below if statements make sure all data is validated based on the format given in the website
        #this ensures no malicious qr codes are scanned
        if(validateFirstName(qr_data[0])):
            self.data_fields["first_name"] = qr_data[0]
        if(validateLastName(qr_data[1])):
            self.data_fields["last_name"] = qr_data[1]
        if(validateRank(qr_data[2])):
            self.data_fields["rank"] = qr_data[2]
        if(validateUnit(qr_data[3])):
            self.data_fields["unit"] = qr_data[3]
        if(validatePhoneNumber(qr_data[4])):
            self.data_fields["phone_number"] = qr_data[4]
        if(validateFitness(qr_data[5])):
            self.data_fields["fitness"] = qr_data[5]
        if(validateProfile(qr_data[6])):
            self.data_fields["profile"] = qr_data[6]

        #if there is a value not found in the data_fields dict, return false
        for key, value in self.data_fields.items():
            if(value is False):
                print(f"No value found for {key}")
                return False
        return True
    
    '''
        Starts loading screen to indicate the machine is working
    '''
    def load_screen(self):
        if self.loading_container.isHidden():
            self.loading_movie = QMovie("images/loading.gif")
            self.loading_label.setMovie(self.loading_movie)
            self.loading_movie.start()
            self.loading_container.show()
            self.loading_container.raise_()
        else:
            self.loading_movie.stop()
            self.loading_container.hide()

    '''
        Shows checkmark animation to indicate a successful scan.
    '''
    def show_checkmark_overlay(self):
        gif_label = QLabel(self)
        movie = QMovie("images/checkmark.gif")

        movie.loopCount = 1
        gif_label.setMovie(movie)
        movie.start()

        screen_geometry = QApplication.desktop().screenGeometry()
        scale_factor = 0.5
        gif_width = int(movie.frameRect().width() * scale_factor)
        gif_height = int(movie.frameRect().height() * scale_factor)
        gif_label.setGeometry(
            (screen_geometry.width() - gif_width) // 2,
            (screen_geometry.height() - gif_height) // 2,
            gif_width,
            gif_height
        )

        # Set aspect ratio policy to keep the GIF aspect ratio when scaled
        gif_label.setScaledContents(True)

        # Connect a function to check if the movie has finished playing
        def check_movie_finished(frame_number):
            if frame_number == movie.frameCount() - 1:
                movie.stop()
                gif_label.close()

        # Connect the function to the frameChanged signal
        movie.frameChanged.connect(check_movie_finished)

        gif_label.show()

    '''
        Below function shows the error messages when a problem is encountered.
    '''

    def show_invalid_qr_message(self, duration=3000):  
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Warning)
        msg.setText("Invalid QR Code.")
        msg.setInformativeText("Please scan again.")
        msg.setWindowTitle("Error")
        
        msg.setWindowFlags(Qt.CustomizeWindowHint | Qt.WindowTitleHint)
        
        timer = QTimer()
        timer.setSingleShot(True)
        timer.timeout.connect(msg.close)
        timer.start(duration)
        
        msg.exec_()
        
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = QRCodeScanner()
    window.show()
    sys.exit(app.exec_())